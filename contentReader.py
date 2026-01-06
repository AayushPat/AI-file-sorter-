"""
Content Reader - Read file contents based on user configuration
Supports text files, PDFs, Office docs, images, and archives
"""
import os
from pathlib import Path
from typing import Dict, Optional, List
import json


def read_text_file(file_path: Path, max_size: int = 5 * 1024 * 1024) -> Optional[str]:
    """Read a text file, respecting max size limit."""
    try:
        if file_path.stat().st_size > max_size:
            return None  # File too large
        
        # Try to read as text
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        return content
    except Exception as e:
        return None


def read_pdf(file_path: Path, max_size: int = 5 * 1024 * 1024) -> Optional[str]:
    """Read text content from a PDF file."""
    try:
        if file_path.stat().st_size > max_size:
            return None
        
        # Try PyPDF2 first
        try:
            import PyPDF2
            with open(file_path, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                text_parts = []
                # Limit to first 10 pages to avoid huge content
                for page_num, page in enumerate(pdf_reader.pages[:10]):
                    try:
                        text = page.extract_text()
                        if text:
                            text_parts.append(text)
                    except:
                        continue
                return "\n".join(text_parts) if text_parts else None
        except ImportError:
            # Try pdfplumber as fallback
            try:
                import pdfplumber
                with pdfplumber.open(file_path) as pdf:
                    text_parts = []
                    for page in pdf.pages[:10]:  # First 10 pages
                        try:
                            text = page.extract_text()
                            if text:
                                text_parts.append(text)
                        except:
                            continue
                    return "\n".join(text_parts) if text_parts else None
            except ImportError:
                return None
    except Exception:
        return None


def read_docx(file_path: Path, max_size: int = 5 * 1024 * 1024) -> Optional[str]:
    """Read text content from a Word document."""
    try:
        if file_path.stat().st_size > max_size:
            return None
        
        try:
            from docx import Document
            doc = Document(file_path)
            paragraphs = [para.text for para in doc.paragraphs]
            return "\n".join(paragraphs)
        except ImportError:
            return None
    except Exception:
        return None


def read_xlsx(file_path: Path, max_size: int = 5 * 1024 * 1024) -> Optional[str]:
    """Read text content from an Excel file."""
    try:
        if file_path.stat().st_size > max_size:
            return None
        
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            text_parts = []
            # Read first sheet only, first 100 rows
            if workbook.sheetnames:
                sheet = workbook[workbook.sheetnames[0]]
                for row in sheet.iter_rows(max_row=100, values_only=True):
                    row_text = " ".join(str(cell) for cell in row if cell)
                    if row_text:
                        text_parts.append(row_text)
            return "\n".join(text_parts) if text_parts else None
        except ImportError:
            return None
    except Exception:
        return None


def read_image_metadata(file_path: Path) -> Optional[Dict]:
    """Extract metadata from an image file."""
    try:
        try:
            from PIL import Image
            from PIL.ExifTags import TAGS
            
            with Image.open(file_path) as img:
                metadata = {
                    "format": img.format,
                    "mode": img.mode,
                    "size": img.size,  # (width, height)
                    "width": img.width,
                    "height": img.height,
                }
                
                # Try to get EXIF data
                try:
                    exif = img._getexif()
                    if exif:
                        exif_data = {}
                        for tag_id, value in exif.items():
                            tag = TAGS.get(tag_id, tag_id)
                            exif_data[tag] = value
                        metadata["exif"] = exif_data
                except:
                    pass
                
                return metadata
        except ImportError:
            return None
    except Exception:
        return None


def read_archive_contents(file_path: Path) -> Optional[List[str]]:
    """List contents of an archive file."""
    try:
        import zipfile
        import tarfile
        
        contents = []
        
        if file_path.suffix.lower() == '.zip':
            try:
                with zipfile.ZipFile(file_path, 'r') as zip_ref:
                    contents = zip_ref.namelist()[:50]  # First 50 files
            except:
                pass
        elif file_path.suffix.lower() in ['.tar', '.tar.gz', '.tgz']:
            try:
                with tarfile.open(file_path, 'r:*') as tar_ref:
                    contents = [name for name in tar_ref.getnames()[:50]]
            except:
                pass
        
        return contents if contents else None
    except Exception:
        return None


def read_file_content(file_path: Path, file_type: str, config: Dict) -> Optional[Dict]:
    """
    Read file content based on file type and user configuration.
    Tries to read ANY file as text if content reading is enabled, even if extension isn't recognized.
    
    Args:
        file_path: Path to the file
        file_type: File extension (e.g., ".pdf", ".txt")
        config: User configuration dict with:
            - enabled_types: List of enabled file type categories
            - max_file_size: Maximum file size to read (in bytes)
            - enabled: Master toggle for content reading
    
    Returns:
        Dict with "content" key containing the read content, or None if not readable/configured
    """
    if not config.get("enabled", False):
        return None
    
    enabled_types = config.get("enabled_types", [])
    max_size = config.get("max_file_size", 5 * 1024 * 1024)  # Default 5MB
    
    file_type_lower = file_type.lower()
    result = {}
    
    # Known binary file types that we shouldn't try to read as text
    binary_extensions = {'.exe', '.dll', '.so', '.dylib', '.bin', '.dat', '.db', '.sqlite', 
                         '.sqlite3', '.db3', '.mdb', '.accdb', '.psd', '.ai', '.sketch',
                         '.dmg', '.iso', '.img', '.deb', '.rpm', '.pkg', '.apk', '.ipa',
                         '.app', '.framework', '.bundle', '.kext', '.dylib', '.a', '.lib',
                         '.o', '.obj', '.class', '.pyc', '.pyo', '.pyd', '.so', '.dll',
                         '.woff', '.woff2', '.ttf', '.otf', '.eot', '.mp3', '.mp4', '.avi',
                         '.mov', '.wmv', '.flv', '.mkv', '.webm', '.m4a', '.wav', '.flac',
                         '.ogg', '.aac', '.wma', '.zip', '.rar', '.7z', '.tar', '.gz',
                         '.bz2', '.xz', '.z', '.lz', '.lzma', '.cab', '.msi', '.pkg',
                         '.dmg', '.iso', '.img', '.vmdk', '.vdi', '.vhd', '.vhdx'}
    
    # PDFs
    if "pdf" in enabled_types and file_type_lower == '.pdf':
        content = read_pdf(file_path, max_size)
        if content is not None:
            result["content"] = content
            result["type"] = "pdf"
            return result
    
    # Office documents
    if "office" in enabled_types:
        if file_type_lower == '.docx':
            content = read_docx(file_path, max_size)
            if content is not None:
                result["content"] = content
                result["type"] = "docx"
                return result
        elif file_type_lower in ['.xlsx', '.xls']:
            content = read_xlsx(file_path, max_size)
            if content is not None:
                result["content"] = content
                result["type"] = "xlsx"
                return result
    
    # Images (metadata only)
    if "images" in enabled_types and file_type_lower in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', 
                                                          '.tiff', '.tif', '.webp', '.svg']:
        metadata = read_image_metadata(file_path)
        if metadata is not None:
            result["metadata"] = metadata
            result["type"] = "image"
            return result
    
    # Archives
    if "archives" in enabled_types and file_type_lower in ['.zip', '.tar', '.tar.gz', '.tgz', 
                                                            '.rar', '.7z']:
        contents = read_archive_contents(file_path)
        if contents is not None:
            result["archive_contents"] = contents
            result["type"] = "archive"
            return result
    
    # Try to read as text for any file type (if text reading is enabled)
    # Skip known binary files, but try to read everything else as text
    if "text" in enabled_types and file_type_lower not in binary_extensions:
        content = read_text_file(file_path, max_size)
        if content is not None:
            result["content"] = content
            result["type"] = "text"
            return result
    
    # Last resort: if content reading is enabled but text reading wasn't enabled,
    # still try reading as text for unknown file types (unless it's a known binary type)
    # This ensures we at least try to read files even if they're not in the enabled types list
    if file_type_lower not in binary_extensions:
        content = read_text_file(file_path, max_size)
        if content is not None:
            result["content"] = content
            result["type"] = "text"
            return result
    
    return None

